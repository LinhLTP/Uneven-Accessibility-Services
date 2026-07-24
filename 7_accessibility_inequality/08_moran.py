# Global Moran's I + Local Moran's I (LISA)

import re
from pathlib import Path

import geopandas as gpd
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from esda.moran import Moran, Moran_Local
from libpysal.weights import KNN

# 1. Config ----
config = {
    "h3_file": Path("accessibility_hanoi/output/accessibility_r5r/accessibility_outputs.gpkg"),
    "h3_layer": "h3_centroids_accessibility",
    "output_dir": Path("accessibility_hanoi/output/accessibility_r5r/moran_outputs"),
    "cutoffs": [15, 30, 45, 60, 90, 120],
    "opp_cols": [
        "healthcare", "education", "retail_grocery", "parks_public_open_space",
        "sports_facilities", "cultural_facilities", "religious_sites",
        "local_government", "food_beverage",
    ],
    "k_neighbors": 6,
    "n_simulations": 499,  # 499 is enough for p < 0.05; use 999 for publication
    "alpha": 0.05,
    "dpi": 300,
}
config["output_dir"].mkdir(parents=True, exist_ok=True)

boundary_district_path = Path("accessibility_hanoi/data_raw/odm_vietnam_district.geojson")

# 2. Palette ----
LISA_COLORS = {
    "High-High": "#D7191C", "Low-Low": "#2C7BB6", "High-Low": "#FDAE61",
    "Low-High": "#ABD9E9", "Not significant": "#F0F0F0",
}
POI_COLORS = {
    "healthcare": "#0072B2", "education": "#56B4E9", "retail_grocery": "#F0E442",
    "parks_public_open_space": "#009E73", "sports_facilities": "#CC79A7",
    "cultural_facilities": "#E69F00", "religious_sites": "#000000",
    "local_government": "#D55E00", "food_beverage": "#999999",
}

# 3. Read H3 layer and reshape to long format ----
# The pipeline outputs a WIDE gpkg (one row per origin, columns like
# healthcare_cum30m, education_cum60m, ...). Pivot to long here: one row
# per origin x opportunity x cutoff.
origins_sf_raw = gpd.read_file(config["h3_file"], layer=config["h3_layer"])

opp_pattern = "^(" + "|".join(config["opp_cols"]) + r")_cum\d+m$"
acc_cols = [c for c in origins_sf_raw.columns if re.match(opp_pattern, c)]
if len(acc_cols) == 0:
    raise ValueError(
        f"No accessibility columns found matching pattern <opp>_cum<cutoff>m.\n"
        f"Available columns: {', '.join(origins_sf_raw.columns)}"
    )

if "origin_id" in origins_sf_raw.columns:
    origins_sf = origins_sf_raw[["origin_id", "geometry"]]
else:
    origins_sf = origins_sf_raw.rename(columns={"id": "origin_id"})[["origin_id", "geometry"]]

id_col_for_long = "origin_id" if "origin_id" in origins_sf_raw.columns else "id"
access_sf = origins_sf_raw.drop(columns="geometry").rename(columns={id_col_for_long: "origin_id"})
access_sf = access_sf[["origin_id"] + acc_cols].melt(
    id_vars="origin_id", value_vars=acc_cols, var_name="col_name", value_name="accessibility"
)
access_sf["opportunity"] = access_sf["col_name"].str.replace(r"_cum\d+m$", "", regex=True)
access_sf["cutoff_min"] = access_sf["col_name"].str.extract(r"cum(\d+)").astype(int)
access_sf["cutoff_label"] = access_sf["cutoff_min"].astype(str) + " min"
access_sf = access_sf[["origin_id", "opportunity", "cutoff_min", "cutoff_label", "accessibility"]]

# 4. Build KNN spatial weights - computed once on the unique origins ----
origins_sf = origins_sf.drop_duplicates(subset="origin_id").sort_values("origin_id").reset_index(drop=True)

coords = np.column_stack([origins_sf.geometry.centroid.x, origins_sf.geometry.centroid.y])
knn_w = KNN.from_array(coords, k=config["k_neighbors"])
knn_w.transform = "r"  # row-standardised weights, mirrors spdep style = "W"

# 5. Global Moran's I - all 54 scenarios ----
scenario_keys = access_sf[["opportunity", "cutoff_min", "cutoff_label"]].drop_duplicates().sort_values(
    ["opportunity", "cutoff_min"]
)

origins_df = origins_sf.drop(columns="geometry")

global_rows = []
for _, key in scenario_keys.iterrows():
    sc_data = access_sf[(access_sf["opportunity"] == key["opportunity"]) & (access_sf["cutoff_min"] == key["cutoff_min"])]
    vals = origins_df.merge(sc_data[["origin_id", "accessibility"]], on="origin_id", how="left")["accessibility"]
    vals = vals.fillna(0).to_numpy()

    mt = Moran(vals, knn_w, permutations=config["n_simulations"])

    global_rows.append({
        "opportunity": key["opportunity"], "cutoff_min": key["cutoff_min"], "cutoff_label": key["cutoff_label"],
        "moran_I": mt.I, "expected_I": mt.EI, "variance_I": mt.VI_rand,
        "z_score": mt.z_rand, "p_value_analytical": mt.p_rand, "p_value_mc": mt.p_sim,
        "n_cells": len(vals), "n_nonzero": int((vals > 0).sum()),
        "significant": mt.p_sim < config["alpha"],
    })

global_moran = pd.DataFrame(global_rows)

# 6. Local Moran's I (LISA) - all 54 scenarios ----
local_rows = []
for _, key in scenario_keys.iterrows():
    sc_data = access_sf[(access_sf["opportunity"] == key["opportunity"]) & (access_sf["cutoff_min"] == key["cutoff_min"])]
    vals = origins_df.merge(sc_data[["origin_id", "accessibility"]], on="origin_id", how="left")["accessibility"]
    vals = vals.fillna(0).to_numpy()

    lm_res = Moran_Local(vals, knn_w, permutations=config["n_simulations"])

    lag_vals = knn_w.sparse @ vals
    z = (vals - vals.mean()) / vals.std(ddof=0)
    lag_z = (lag_vals - lag_vals.mean()) / lag_vals.std(ddof=0)
    p_val = lm_res.p_sim

    quad = np.select(
        [p_val >= config["alpha"], (z > 0) & (lag_z > 0), (z < 0) & (lag_z < 0),
         (z > 0) & (lag_z < 0), (z < 0) & (lag_z > 0)],
        ["Not significant", "High-High", "Low-Low", "High-Low", "Low-High"],
        default="Not significant",
    )

    local_rows.append(pd.DataFrame({
        "origin_id": origins_df["origin_id"], "opportunity": key["opportunity"],
        "cutoff_min": key["cutoff_min"], "cutoff_label": key["cutoff_label"],
        "accessibility": vals, "local_I": lm_res.Is, "z_score_local": lm_res.z_sim,
        "p_value_local": p_val, "lag_access": lag_vals, "lisa_cluster": quad,
    }))

local_moran_all = pd.concat(local_rows, ignore_index=True)

# 7. Export results to Excel ----
out_xlsx = config["output_dir"] / "moran_results.xlsx"

lisa_summary = (
    local_moran_all.groupby(["opportunity", "cutoff_min", "cutoff_label", "lisa_cluster"])
    .size().reset_index(name="n_cells")
    .pivot_table(index=["opportunity", "cutoff_min", "cutoff_label"], columns="lisa_cluster",
                 values="n_cells", fill_value=0)
    .reset_index()
)

with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
    global_moran.to_excel(writer, sheet_name="Global_Moran", index=False)
    lisa_summary.to_excel(writer, sheet_name="LISA_Summary", index=False)

    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    header_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    header_border = Border(bottom=Side(style="thin"))
    for sheet_name, df in [("Global_Moran", global_moran), ("LISA_Summary", lisa_summary)]:
        ws = writer.sheets[sheet_name]
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = header_border
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        for j, col_name in enumerate(df.columns, start=1):
            width = max(14, len(str(col_name)) + 3)
            ws.column_dimensions[ws.cell(row=1, column=j).column_letter].width = width
        ws.freeze_panes = "A2"

# 8. Plots ----

# 8.1. Text / theme config
TEXT_CONFIG = {"base_size": 11, "title_size": 13, "axis_size": 10, "legend_size": 10,
                "strip_size": 9, "caption_size": 8}

# 8.2. Global Moran's I heatmap
poi_order = global_moran.groupby("opportunity")["moran_I"].mean().sort_values(ascending=False).index.tolist()
cutoff_levels = [f"{c} min" for c in sorted(config["cutoffs"])]

heat = global_moran.copy()
heat["opportunity"] = pd.Categorical(heat["opportunity"], categories=poi_order, ordered=True)
heat["cutoff_label"] = pd.Categorical(heat["cutoff_label"], categories=cutoff_levels, ordered=True)
heat["sig_label"] = np.where(heat["significant"], "*", "")

pivot_I = heat.pivot(index="opportunity", columns="cutoff_label", values="moran_I").reindex(index=poi_order[::-1])
pivot_sig = heat.pivot(index="opportunity", columns="cutoff_label", values="sig_label").reindex(index=poi_order[::-1])

fig, ax = plt.subplots(figsize=(10, 6))
im = ax.imshow(pivot_I.to_numpy(), cmap="cividis", vmin=0, vmax=1, aspect="auto")
ax.set_xticks(range(len(pivot_I.columns)))
ax.set_xticklabels(pivot_I.columns)
ax.xaxis.tick_top()
ax.set_yticks(range(len(pivot_I.index)))
ax.set_yticklabels(pivot_I.index)
for i in range(pivot_I.shape[0]):
    for j in range(pivot_I.shape[1]):
        val = pivot_I.iloc[i, j]
        if pd.notna(val):
            ax.text(j, i, f"{val:.3f}{pivot_sig.iloc[i, j]}", ha="center", va="center", fontsize=8.5)
fig.colorbar(im, ax=ax, label="Moran's I")
fig.text(0.5, 0.02,
         "Global Moran's I by POI type and travel time threshold\n"
         "(* = significant at p < 0.05, Monte Carlo permutation, n=" + str(config["n_simulations"] + 1) + ")",
         ha="center", fontsize=TEXT_CONFIG["axis_size"])
p_global = fig

# 8.3. Moran scatter plots - one panel per POI at a focal cutoff
focal_cutoff = 30

scatter_figs = []
for opp in poi_order:
    d = local_moran_all[
        (local_moran_all["opportunity"] == opp) & (local_moran_all["cutoff_min"] == focal_cutoff)
        & (local_moran_all["accessibility"] > 0)
    ]
    if len(d) < 10:
        continue

    z = (d["accessibility"] - d["accessibility"].mean()) / d["accessibility"].std(ddof=0)
    lag_z = (d["lag_access"] - d["lag_access"].mean()) / d["lag_access"].std(ddof=0)
    global_I = global_moran.loc[
        (global_moran["opportunity"] == opp) & (global_moran["cutoff_min"] == focal_cutoff), "moran_I"
    ].iloc[0]

    scatter_figs.append((opp, z, lag_z, d["lisa_cluster"], global_I, len(d)))

n_scatter = len(scatter_figs)
ncol_scatter = 4
nrow_scatter = int(np.ceil(n_scatter / ncol_scatter))
fig_s, axes_s = plt.subplots(nrow_scatter, ncol_scatter, figsize=(16, 10))
axes_s = np.atleast_1d(axes_s).flatten()

for ax, (opp, z, lag_z, cluster, global_I, n_pos) in zip(axes_s, scatter_figs):
    ax.axhline(0, linestyle="--", color="#999999")
    ax.axvline(0, linestyle="--", color="#999999")
    colours = cluster.map(LISA_COLORS)
    ax.scatter(z, lag_z, c=colours, alpha=0.5, s=8)
    slope, intercept = np.polyfit(z, lag_z, 1)
    xs = np.linspace(z.min(), z.max(), 50)
    ax.plot(xs, slope * xs + intercept, color="black", linewidth=0.8)
    ax.set_title(opp, fontsize=TEXT_CONFIG["axis_size"], fontweight="bold")
    ax.text(0.02, 0.98, f"I = {global_I:.3f}  (n = {n_pos} positive cells)",
            transform=ax.transAxes, ha="left", va="top", fontsize=TEXT_CONFIG["caption_size"], color="#666666")
    ax.set_xlabel("Standardised accessibility (z)", fontsize=8)
    ax.set_ylabel("Standardised spatial lag (z)", fontsize=8)

for ax in axes_s[n_scatter:]:
    ax.set_visible(False)

fig_s.suptitle(f"Moran Scatter Plots \u2014 {focal_cutoff} min threshold", fontsize=TEXT_CONFIG["title_size"], fontweight="bold")
fig_s.text(0.5, 0.965,
           "Each point = one H3 cell with accessibility > 0. Slope = Global Moran's I. Colour = LISA quadrant.",
           ha="center", fontsize=TEXT_CONFIG["axis_size"], color="#4d4d4d")
fig_s.text(0.02, 0.005,
           "Note: zero-accessibility cells excluded to show all four quadrants clearly.\n"
           "Source: r5py routing, Hanoi 2018. KNN k=6 spatial weights.",
           fontsize=TEXT_CONFIG["caption_size"], color="#808080")
p_scatter = fig_s

# 8.4. LISA cluster maps - facet by cutoff, one file per POI
hanoi_boundary = gpd.read_file(boundary_district_path)
hanoi_boundary["geometry"] = hanoi_boundary.geometry.make_valid()

def _std(x):
    return x.astype(str).map(lambda s: "".join(c for c in s.lower() if c.isalnum() or c == " ")).str.strip()

hanoi_boundary = hanoi_boundary[_std(hanoi_boundary["Province"]).isin(["ha noi", "hanoi"])]
hanoi_boundary = hanoi_boundary.to_crs(origins_sf.crs)

origins_geom = origins_sf[["origin_id", "geometry"]]

for opp in poi_order:
    lisa_sf = local_moran_all[local_moran_all["opportunity"] == opp].copy()
    lisa_sf["cutoff_label"] = pd.Categorical(lisa_sf["cutoff_label"], categories=cutoff_levels, ordered=True)
    lisa_sf["lisa_cluster"] = pd.Categorical(lisa_sf["lisa_cluster"], categories=list(LISA_COLORS.keys()), ordered=True)
    lisa_sf = lisa_sf.merge(origins_geom, on="origin_id", how="left")
    lisa_sf = gpd.GeoDataFrame(lisa_sf, geometry="geometry", crs=origins_sf.crs)

    fig_m, axes_m = plt.subplots(1, len(cutoff_levels), figsize=(22, 5))
    for ax, cutoff_label in zip(axes_m, cutoff_levels):
        hanoi_boundary.plot(ax=ax, facecolor="#f5f5f5", edgecolor="#999999", linewidth=0.35)
        subset = lisa_sf[lisa_sf["cutoff_label"] == cutoff_label]
        for cluster, colour in LISA_COLORS.items():
            pts = subset[subset["lisa_cluster"] == cluster]
            if len(pts) > 0:
                pts.plot(ax=ax, color=colour, markersize=0.25, alpha=0.85, label=cluster)
        hanoi_boundary.plot(ax=ax, facecolor="none", edgecolor="#666666", linewidth=0.25)
        ax.set_title(cutoff_label, fontsize=TEXT_CONFIG["strip_size"], fontweight="bold")
        ax.set_xticks([])
        ax.set_yticks([])

    handles = [plt.Line2D([0], [0], marker="o", color="w", markerfacecolor=c, label=k, markersize=8)
               for k, c in LISA_COLORS.items()]
    fig_m.legend(handles=handles, title="LISA cluster", loc="lower center", ncol=5, fontsize=8)
    fig_m.suptitle(f"LISA Cluster Map \u2014 {opp}", fontsize=TEXT_CONFIG["title_size"], fontweight="bold")
    fig_m.text(0.5, 0.93, "H3 hexagon centroids, KNN k=6 weights, p < 0.05 (two-sided)",
               ha="center", fontsize=TEXT_CONFIG["axis_size"], color="#4d4d4d")
    fig_m.text(0.02, 0.01,
               "HH = High-High (hot spot)  LL = Low-Low (cold spot)\n"
               "HL/LH = spatial outliers  Grey = not significant\n"
               "District boundaries shown for reference.\nSource: r5py routing, Hanoi 2018.",
               fontsize=TEXT_CONFIG["caption_size"], color="#808080")

    fname = config["output_dir"] / f"LISA_map_{opp.replace(' ', '_')}.png"
    fig_m.savefig(fname, dpi=config["dpi"], bbox_inches="tight", facecolor="white")
    plt.close(fig_m)

# 8.5. Save global heatmap + scatter
p_global.savefig(config["output_dir"] / "global_moran_heatmap.png", dpi=config["dpi"], bbox_inches="tight", facecolor="white")
plt.close(p_global)

p_scatter.savefig(config["output_dir"] / f"moran_scatter_{focal_cutoff}min.png", dpi=config["dpi"], bbox_inches="tight", facecolor="white")
plt.close(p_scatter)
