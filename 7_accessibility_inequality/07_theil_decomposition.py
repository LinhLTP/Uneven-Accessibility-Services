import unicodedata
from pathlib import Path

import geopandas as gpd
import numpy as np
import pandas as pd
import requests

# 1. Config ----
config = {
    "h3_file": Path("accessibility_hanoi/output/accessibility_r5r/accessibility_outputs.gpkg"),
    "h3_layer": "h3_centroids_accessibility",
    "acc_long_csv": Path("accessibility_hanoi/output/accessibility_r5r/accessibility_long.csv"),
    "boundary_url": (
        "https://data.opendevelopmentmekong.net/dataset/6f054351-bf2c-422e-8deb-0a511d63a315/"
        "resource/c906af7a-e7a0-4776-95d4-5ee815dba760/download/district.geojson"
    ),
    "boundary_local": Path("accessibility_hanoi/data_raw/odm_vietnam_district.geojson"),
    "output_dir": Path("accessibility_hanoi/output/accessibility_r5r/theil_urban_rural_outputs"),
    # Theil T is computed on accessibility > 0 cells only (zeros excluded).
    # See pct_zero_excluded in the output to report the exclusion rate.
}

config["boundary_local"].parent.mkdir(parents=True, exist_ok=True)
config["output_dir"].mkdir(parents=True, exist_ok=True)

# 2. Helper functions ----

# 2.1. Name standardisation
def standardise_name(x):
    x = x.astype(str).map(lambda s: unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode())
    return x.str.lower().str.replace(r"[^a-z0-9]+", " ", regex=True).str.strip()


# 2.2. Guess province/district column names
def guess_admin_cols(gdf):
    nm = list(gdf.columns)
    nm_std = standardise_name(pd.Series(nm)).tolist()

    province_candidates = ["province", "province name", "provincename", "name 1", "name1", "adm1", "admin1", "tinh", "ten tinh"]
    district_candidates = ["district", "district name", "districtname", "name 2", "name2", "adm2", "admin2", "huyen", "ten huyen", "quan huyen"]

    def find_first_match(candidates):
        for i, n in enumerate(nm_std):
            if n in candidates:
                return nm[i]
        for term in candidates:
            for i, n in enumerate(nm_std):
                if term in n:
                    return nm[i]
        return None

    province_col = find_first_match(province_candidates) or ("NAME_1" if "NAME_1" in nm else None) or ("Province" if "Province" in nm else None)
    district_col = find_first_match(district_candidates) or ("NAME_2" if "NAME_2" in nm else None) or ("District" if "District" in nm else None)
    return province_col, district_col


# 2.3. Download boundary if needed
def download_boundary_if_needed(url, destfile):
    if not destfile.exists():
        with requests.get(url, stream=True, timeout=120) as r:
            r.raise_for_status()
            with open(destfile, "wb") as f:
                for chunk in r.iter_content(chunk_size=8192):
                    f.write(chunk)


# 2.4. Theil T decomposition
# NOTE ON TERMINOLOGY (Theil T adapted for accessibility):
#   "population" in classic Theil = number of H3 cells here
#   "income"     in classic Theil = cumulative accessibility of a group
#   No demographic or income data is used.
#
# Formula:
#   T_total   = theil_t(all positive accessibility values)
#   T_between = sum_g [ access_share_g * ln(mean_ratio_g) ]
#   T_within  = sum_g [ access_share_g * T_g ]
#   T_total   = T_between + T_within  (additive decomposition)
def theil_t(x):
    x = np.asarray(x, dtype=float)
    x = x[x > 0]
    if len(x) == 0:
        return np.nan
    mu = x.mean()
    if mu == 0:
        return np.nan
    return float(np.mean((x / mu) * np.log(x / mu)))


def theil_decomposition(vals, groups):
    vals = np.asarray(vals, dtype=float)
    groups = np.asarray(groups)
    keep = np.isfinite(vals) & (vals > 0)
    vals_pos, grps_pos = vals[keep], groups[keep]

    n = len(vals_pos)
    mu = vals_pos.mean() if n > 0 else np.nan

    if n == 0 or mu == 0:
        return {
            "overall": pd.DataFrame([{
                "n_cells_total": n, "mean_access": mu, "theil_total": np.nan,
                "theil_between": np.nan, "theil_within": np.nan,
                "between_share": np.nan, "within_share": np.nan, "decomposition_gap": np.nan,
            }]),
            "details": pd.DataFrame(),
        }

    T_total = theil_t(vals_pos)

    rows = []
    for g in sorted(pd.unique(grps_pos)):
        v = vals_pos[grps_pos == g]
        n_g = len(v)
        mu_g = v.mean()
        T_g = theil_t(v) if n_g > 1 else 0.0

        acc_s = (n_g * mu_g) / (n * mu)
        m_ratio = mu_g / mu
        bet_c = acc_s * np.log(m_ratio) if mu_g > 0 else 0.0
        wit_c = acc_s * T_g

        rows.append({
            "group_name": g, "n_cells_group": n_g, "mean_access_group": mu_g,
            "theil_within_group": T_g, "cell_share": n_g / n, "access_share": acc_s,
            "mean_ratio": m_ratio, "between_component": bet_c, "within_component": wit_c,
        })

    details = pd.DataFrame(rows)
    T_between = details["between_component"].sum()
    T_within = details["within_component"].sum()

    overall = pd.DataFrame([{
        "n_cells_total": n, "mean_access": mu, "theil_total": T_total,
        "theil_between": T_between, "theil_within": T_within,
        "between_share": T_between / T_total if T_total not in (0, np.nan) else np.nan,
        "within_share": T_within / T_total if T_total not in (0, np.nan) else np.nan,
        "decomposition_gap": T_total - T_between - T_within,
    }])

    return {"overall": overall, "details": details}


# 2.5. Assign points to districts (within, falling back to nearest)
def assign_points_to_districts(points_gdf, districts_gdf):
    districts_gdf = districts_gdf.to_crs(points_gdf.crs)

    within = gpd.sjoin(points_gdf, districts_gdf, predicate="within", how="left")
    within = within[~within.index.duplicated(keep="first")]
    within["assignment_method"] = np.where(within.index_right.notna(), "within", "nearest_fallback")

    missing = within["index_right"].isna()
    if missing.any():
        nearest = gpd.sjoin_nearest(points_gdf[missing], districts_gdf)
        nearest = nearest[~nearest.index.duplicated(keep="first")]
        for col in districts_gdf.columns:
            if col != "geometry":
                within.loc[missing, col] = nearest[col].to_numpy()

    return within.drop(columns=[c for c in ["index_right"] if c in within.columns])


# 3. Boundary and district classification ----

# 3.1. Download and read OD Mekong district boundary
download_boundary_if_needed(config["boundary_url"], config["boundary_local"])

odm_dist = gpd.read_file(config["boundary_local"])
odm_dist["geometry"] = odm_dist.geometry.make_valid()
province_col, district_col = guess_admin_cols(odm_dist)
if province_col is None or district_col is None:
    raise ValueError(f"Could not detect province/district columns. Available: {list(odm_dist.columns)}")

odm_dist["province_std"] = standardise_name(odm_dist[province_col].astype(str))
odm_dist["district_std"] = standardise_name(odm_dist[district_col].astype(str))

hanoi_districts = odm_dist[odm_dist["province_std"].isin(["ha noi", "hanoi"])].copy()
hanoi_districts = hanoi_districts.rename(columns={province_col: "province_name", district_col: "district_name"})
hanoi_districts = hanoi_districts[["province_name", "district_name", "province_std", "district_std", "geometry"]]
hanoi_districts["geometry"] = hanoi_districts.geometry.make_valid()

if len(hanoi_districts) == 0:
    raise ValueError("No Hanoi districts found.")

# 3.2. Classify districts into Urban / Rural
urban_districts = [
    "Hoan Kiem", "Ba Dinh", "Dong Da", "Hai Ba Trung", "Thanh Xuan", "Cau Giay",
    "Tay Ho", "Hoang Mai", "Long Bien", "Ha Dong", "Tu Liem", "Bac Tu Liem", "Nam Tu Liem",
]
rural_districts = [
    "Ba Vi", "Soc Son", "My Duc", "Chuong My", "Dong Anh", "Ung Hoa", "Phu Xuyen",
    "Thach That", "Quoc Oai", "Thuong Tin", "Me Linh", "Thanh Oai", "Son Tay",
    "Phuc Tho", "Gia Lam", "Hoai Duc", "Dan Phuong", "Thanh Tri",
]
urban_std = set(standardise_name(pd.Series(urban_districts)))
rural_std = set(standardise_name(pd.Series(rural_districts)))

hanoi_districts["urban_rural_group"] = np.select(
    [hanoi_districts["district_std"].isin(urban_std), hanoi_districts["district_std"].isin(rural_std)],
    ["Urban", "Rural"], default="Unclassified",
)
hanoi_districts["admin_group_detail"] = hanoi_districts["urban_rural_group"].map(
    {"Urban": "Urban District", "Rural": "Town/Rural District"}
).fillna("Unclassified")

unclassified = hanoi_districts.loc[hanoi_districts["urban_rural_group"] == "Unclassified", "district_name"].unique()
if len(unclassified) > 0:
    raise ValueError(f"Some districts unclassified. Please review the district list: {list(unclassified)}")

# 4. Read accessibility data ----
available_layers = gpd.list_layers(config["h3_file"])["name"].tolist()
if config["h3_layer"] not in available_layers:
    raise ValueError(f"Layer not found: {config['h3_layer']}\nAvailable: {', '.join(available_layers)}")

h3_geom_sf = gpd.read_file(config["h3_file"], layer=config["h3_layer"])[["id", "geometry"]]
h3_geom_sf = h3_geom_sf.rename(columns={"id": "origin_id"})
h3_geom_sf["origin_id"] = h3_geom_sf["origin_id"].astype(str)
if h3_geom_sf.crs is None:
    raise ValueError("Accessibility layer has no CRS.")

acc_long_raw = pd.read_csv(config["acc_long_csv"])
if "origin_id" not in acc_long_raw.columns and "id" in acc_long_raw.columns:
    acc_long_raw = acc_long_raw.rename(columns={"id": "origin_id"})
if "cutoff_label" not in acc_long_raw.columns:
    acc_long_raw["cutoff_label"] = acc_long_raw["cutoff"].astype(str) + " min"
if "cutoff_min" not in acc_long_raw.columns:
    acc_long_raw = acc_long_raw.rename(columns={"cutoff": "cutoff_min"})

required_cols = {"origin_id", "opportunity", "cutoff_min", "accessibility"}
missing_cols = required_cols - set(acc_long_raw.columns)
if missing_cols:
    raise ValueError(f"Missing columns in long CSV: {', '.join(missing_cols)}")

acc_long_raw["origin_id"] = acc_long_raw["origin_id"].astype(str)
access_sf = gpd.GeoDataFrame(
    acc_long_raw.merge(h3_geom_sf, on="origin_id", how="left"), geometry="geometry", crs=h3_geom_sf.crs
)

# 5. Assign H3 origins to Urban / Rural districts ----
origins_sf = access_sf[["origin_id", "geometry"]].drop_duplicates(subset="origin_id")
origins_assigned = assign_points_to_districts(origins_sf, hanoi_districts)

# 6. Join Urban/Rural label back to the accessibility long table ----
assignment_lookup = origins_assigned[
    ["origin_id", "district_name", "urban_rural_group", "admin_group_detail", "assignment_method"]
]

access_long = access_sf.merge(assignment_lookup, on="origin_id", how="left")
if access_long["urban_rural_group"].isna().any():
    raise ValueError("Some rows were not assigned to Urban/Rural.")

access_df = access_long.drop(columns="geometry")

# 7. Theil T - whole-city total per scenario ----
def city_summary(g):
    positive = g.loc[g["accessibility"] > 0, "accessibility"]
    return pd.Series({
        "n_cells_total": len(g),
        "n_zero": (g["accessibility"] == 0).sum(),
        "n_positive": (g["accessibility"] > 0).sum(),
        "pct_zero_excluded": (g["accessibility"] == 0).sum() / len(g),
        "mean_access_all": g["accessibility"].mean(skipna=True),
        "mean_access_pos": positive.mean() if len(positive) > 0 else np.nan,
        "median_access_pos": positive.median() if len(positive) > 0 else np.nan,
        "theil_total": theil_t(positive) if len(positive) > 1 else np.nan,
    })


theil_total_city = (
    access_df.groupby(["opportunity", "cutoff_min", "cutoff_label"])
    .apply(city_summary)
    .reset_index()
)

# 8. Theil decomposition: between Urban/Rural + within each group ----
scenario_keys = access_df[["opportunity", "cutoff_min", "cutoff_label"]].drop_duplicates()

overall_rows, details_rows = [], []
for _, key in scenario_keys.iterrows():
    df_scn = access_df[
        (access_df["opportunity"] == key["opportunity"])
        & (access_df["cutoff_min"] == key["cutoff_min"])
        & (access_df["cutoff_label"] == key["cutoff_label"])
    ]
    if len(df_scn) == 0:
        continue

    dec = theil_decomposition(df_scn["accessibility"], df_scn["urban_rural_group"])

    overall = dec["overall"].copy()
    overall.insert(0, "opportunity", key["opportunity"])
    overall.insert(1, "cutoff_min", key["cutoff_min"])
    overall.insert(2, "cutoff_label", key["cutoff_label"])
    overall_rows.append(overall)

    details = dec["details"].copy()
    if len(details) > 0:
        details.insert(0, "opportunity", key["opportunity"])
        details.insert(1, "cutoff_min", key["cutoff_min"])
        details.insert(2, "cutoff_label", key["cutoff_label"])
        details_rows.append(details)

theil_decomp_overall = pd.concat(overall_rows, ignore_index=True)
theil_decomp_details = pd.concat(details_rows, ignore_index=True)

# 9. Export: 3 sheets in one Excel workbook ----
out_xlsx = config["output_dir"] / "theil_urban_rural_decomposition.xlsx"

with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
    theil_total_city.to_excel(writer, sheet_name="Theil_City_Total", index=False)
    theil_decomp_overall.to_excel(writer, sheet_name="Theil_Decomp_Overall", index=False)
    theil_decomp_details.to_excel(writer, sheet_name="Theil_Decomp_GroupDetails", index=False)

    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    header_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    header_border = Border(bottom=Side(style="thin"))
    for sheet_name, df in [
        ("Theil_City_Total", theil_total_city),
        ("Theil_Decomp_Overall", theil_decomp_overall),
        ("Theil_Decomp_GroupDetails", theil_decomp_details),
    ]:
        ws = writer.sheets[sheet_name]
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = header_border
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        for j, col_name in enumerate(df.columns, start=1):
            width = max(14, len(str(col_name)) + 3)
            ws.column_dimensions[ws.cell(row=1, column=j).column_letter].width = width
        ws.freeze_panes = "A2"
