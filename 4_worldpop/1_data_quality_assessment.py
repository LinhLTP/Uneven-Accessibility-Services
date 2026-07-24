# 1_data_quality_assessment.py

import re
from pathlib import Path

import geopandas as gpd
import numpy as np
import pandas as pd
import rasterio
from esda.moran import Moran
from libpysal.weights import KNN
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# 1. Setup ----

# 1.1. Imports

# 1.2. Config
BASE_DIR = Path("/Volumes/2832083l/02_hanoi_accessibility/accessibility_hn")

CFG = {
    # Raster paths (from worldpop_01)
    "clip_dir": BASE_DIR / "data/worldpop_hanoi_2018/clipped",
    "total_pop_tif": BASE_DIR / "data/worldpop_hanoi_2018/clipped/vnm_ppp_2018_hanoi.tif",
    # H3 paths (from worldpop_02)
    "h3_dir": BASE_DIR / "data/worldpop_hanoi_2018/h3_res9",
    "worldpop_total": BASE_DIR / "data/worldpop_hanoi_2018/h3_res9/worldpop_hanoi_h3_res9_total_pop.gpkg",
    "worldpop_age_sex": BASE_DIR / "data/worldpop_hanoi_2018/h3_res9/worldpop_hanoi_h3_res9_age_sex_wide.gpkg",
    "h3_geom_gpkg": BASE_DIR / "data/worldpop_hanoi_2018/h3_res9/worldpop_hanoi_h3_res9_geometry.gpkg",
    "validation_csv": BASE_DIR / "data/worldpop_hanoi_2018/h3_res9/worldpop_hanoi_h3_res9_validation.csv",
    # Accessibility
    "access_wide_csv": BASE_DIR / "output/accessibility_r5r/accessibility_wide.csv",
    # Boundary
    "boundary_gpkg": BASE_DIR / "data/boundaries/hanoi_boundary.gpkg",
    # Output
    "out_excel": BASE_DIR / "output/worldpop_data_quality_assessment.xlsx",
    # Reference: Hanoi population 2018 (GSO / World Bank estimate)
    # Source: Vietnam General Statistics Office, Statistical Yearbook 2018
    "census_ref_pop": 8053663,   # update if a more precise figure becomes available
    "census_ref_year": 2018,
    "knn_k": 6,
    "crs_proj": 32648,
}

CFG["out_excel"].parent.mkdir(parents=True, exist_ok=True)

# 1.3. Excel workbook helpers
wb = Workbook()
wb.remove(wb.active)

HEADER_STYLE_A = {"fill": "1F3864", "font": "FFFFFF"}
HEADER_STYLE_B = {"fill": "375623", "font": "FFFFFF"}


def add_sheet(wb, name, df, header_style=HEADER_STYLE_A):
    ws = wb.create_sheet(name[:31])
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    fill = PatternFill(fill_type="solid", fgColor=header_style["fill"])
    font = Font(color=header_style["font"], bold=True)
    align = Alignment(horizontal="center")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = align

    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 60)


def fmt_int(x):
    return f"{round(x):,}"


# 2. Part A - Raster grid assessment ----

# 2.1. A0. Load boundary
hanoi_gdf = gpd.read_file(CFG["boundary_gpkg"])
hanoi_gdf["geometry"] = hanoi_gdf.geometry.make_valid()
hanoi_gdf = hanoi_gdf.to_crs(4326)
hanoi_vect = [hanoi_gdf.geometry.union_all()]
hanoi_area_km2 = hanoi_gdf.to_crs(CFG["crs_proj"]).geometry.area.sum() / 1e6

# 2.2. A1. Spatial coverage
if CFG["total_pop_tif"].exists():
    with rasterio.open(CFG["total_pop_tif"]) as src:
        r_total_arr = src.read(1).astype("float64")
        r_total_nodata = src.nodata
        r_total_transform = src.transform
        r_total_crs_name = src.crs.to_string() if src.crs else "unknown"

    if r_total_nodata is not None:
        r_total_arr[r_total_arr == r_total_nodata] = np.nan

    n_total_pixels = r_total_arr.size
    n_valid_pixels = int(np.sum(~np.isnan(r_total_arr)))
    n_na_pixels = n_total_pixels - n_valid_pixels
    pixel_res_m = round(abs(r_total_transform.a) * 111320, 1)
    pixel_area_km2 = (abs(r_total_transform.a) * 111320) ** 2 / 1e6

    coverage_df = pd.DataFrame({
        "Metric": [
            "Raster CRS", "Pixel resolution (approx. metres)",
            "Total pixels in extent", "Valid pixels (non-NA)",
            "NoData pixels (NA)", "Coverage rate (%)",
            "Valid area (km\u00b2)", "Hanoi boundary area (km\u00b2)",
            "Coverage vs boundary (%)",
        ],
        "Value": [
            r_total_crs_name,
            f"~{pixel_res_m} m",
            fmt_int(n_total_pixels),
            fmt_int(n_valid_pixels),
            fmt_int(n_na_pixels),
            round(n_valid_pixels / n_total_pixels * 100, 2),
            round(n_valid_pixels * pixel_area_km2, 1),
            round(hanoi_area_km2, 1),
            round(n_valid_pixels * pixel_area_km2 / hanoi_area_km2 * 100, 1),
        ],
    })
else:
    coverage_df = pd.DataFrame({"Metric": ["File not found"], "Value": [str(CFG["total_pop_tif"])]})

add_sheet(wb, "A1_Spatial_Coverage", coverage_df)

# 2.3. A2. Pixel value distribution (total pop)
if CFG["total_pop_tif"].exists():
    vals = r_total_arr[~np.isnan(r_total_arr)]

    dist_df = pd.DataFrame({
        "Metric": [
            "N valid pixels", "Sum (total population)",
            "Mean pop/pixel", "Median pop/pixel",
            "SD", "Min", "P1", "P5", "P25", "P75", "P95", "P99", "Max",
            "% pixels with pop = 0", "% pixels with pop < 1",
            "% pixels with pop >= 10", "% pixels with pop >= 50",
        ],
        "Value": [
            fmt_int(len(vals)),
            fmt_int(round(vals.sum())),
            round(float(np.mean(vals)), 4),
            round(float(np.median(vals)), 4),
            round(float(np.std(vals, ddof=1)), 4),
            round(float(np.min(vals)), 4),
            round(float(np.percentile(vals, 1)), 4),
            round(float(np.percentile(vals, 5)), 4),
            round(float(np.percentile(vals, 25)), 4),
            round(float(np.percentile(vals, 75)), 4),
            round(float(np.percentile(vals, 95)), 4),
            round(float(np.percentile(vals, 99)), 4),
            round(float(np.max(vals)), 4),
            f"{round(float(np.mean(vals == 0)) * 100, 2)}%",
            f"{round(float(np.mean(vals < 1)) * 100, 2)}%",
            f"{round(float(np.mean(vals >= 10)) * 100, 2)}%",
            f"{round(float(np.mean(vals >= 50)) * 100, 2)}%",
        ],
    })
    raster_total_pop = float(vals.sum())
else:
    dist_df = pd.DataFrame({"Metric": ["File not found"], "Value": [str(CFG["total_pop_tif"])]})
    raster_total_pop = np.nan

add_sheet(wb, "A2_Pixel_Distribution", dist_df)

# 2.4. A3. Age-sex rasters: internal consistency
age_tifs = sorted(CFG["clip_dir"].glob("*.tif"))
age_tifs = [f for f in age_tifs if re.match(r"^vnm_(f|m)_\d+_2018_hanoi\.tif$", f.name)]


def age_group_label(code):
    if code == "0":
        return "0\u201311m"
    if code == "1":
        return "1\u20134"
    if code == "80":
        return "80+"
    return f"{code}\u2013{int(code) + 4}"


if len(age_tifs) > 0:
    rows = []
    for f in age_tifs:
        with rasterio.open(f) as src:
            arr = src.read(1).astype("float64")
            nodata = src.nodata
        if nodata is not None:
            arr[arr == nodata] = np.nan
        s = float(np.nansum(arr))
        m = re.match(r"^vnm_(f|m)_(\d+)_2018_hanoi$", f.stem)
        rows.append({
            "File": f.name,
            "Sex": "Female" if m.group(1) == "f" else "Male",
            "Age_code": int(m.group(2)),
            "Age_group": age_group_label(m.group(2)),
            "Sum_population": round(s, 1),
        })
    age_sums = pd.DataFrame(rows).sort_values(["Sex", "Age_code"])

    total_age_sex = age_sums["Sum_population"].sum()
    female_total = age_sums.loc[age_sums["Sex"] == "Female", "Sum_population"].sum()
    male_total = age_sums.loc[age_sums["Sex"] == "Male", "Sum_population"].sum()

    consistency_summary = pd.DataFrame({
        "Metric": [
            "Sum of all age-sex rasters", "Female sub-total", "Male sub-total",
            "Total pop raster sum", "Difference (age-sex sum \u2212 total pop)",
            "Difference (%)", "% Female (from age-sex rasters)",
        ],
        "Value": [
            fmt_int(total_age_sex), fmt_int(female_total), fmt_int(male_total),
            fmt_int(raster_total_pop),
            fmt_int(total_age_sex - raster_total_pop),
            f"{round((total_age_sex - raster_total_pop) / raster_total_pop * 100, 3)}%",
            f"{round(female_total / total_age_sex * 100, 2)}%",
        ],
    })
    consistency_summary["Value"] = consistency_summary["Value"].astype(str)

    a3_bands = pd.concat([
        pd.DataFrame({"Metric": ["--- Female bands ---"], "Value": [None]}),
        age_sums.loc[age_sums["Sex"] == "Female", ["Age_group", "Sum_population"]]
            .rename(columns={"Age_group": "Metric", "Sum_population": "Value"}),
        pd.DataFrame({"Metric": ["--- Male bands ---"], "Value": [None]}),
        age_sums.loc[age_sums["Sex"] == "Male", ["Age_group", "Sum_population"]]
            .rename(columns={"Age_group": "Metric", "Sum_population": "Value"}),
    ], ignore_index=True)
    a3_bands["Value"] = a3_bands["Value"].astype(str)

    add_sheet(wb, "A3_AgeSex_Consistency", pd.concat([consistency_summary, a3_bands], ignore_index=True))
else:
    add_sheet(wb, "A3_AgeSex_Consistency",
              pd.DataFrame({"Note": ["No age-sex rasters found"], "Path": [str(CFG["clip_dir"])]}))

# 2.5. A4. External validation vs census
ext_val_df = pd.DataFrame({
    "Source": [
        "WorldPop 2018 raster (clipped to Hanoi boundary)",
        "WorldPop 2018 - age-sex sum",
        "Reference: GSO Statistical Yearbook 2018",
        "Difference: raster vs census",
        "Difference (%): raster vs census",
    ],
    "Value": [
        fmt_int(raster_total_pop) if not np.isnan(raster_total_pop) else "N/A",
        fmt_int(total_age_sex) if len(age_tifs) > 0 else "N/A",
        fmt_int(CFG["census_ref_pop"]),
        fmt_int(raster_total_pop - CFG["census_ref_pop"]) if not np.isnan(raster_total_pop) else "N/A",
        f"{round((raster_total_pop - CFG['census_ref_pop']) / CFG['census_ref_pop'] * 100, 2)}%"
        if not np.isnan(raster_total_pop) else "N/A",
    ],
    "Note": [
        "Sum of non-NA pixel values",
        "Sum of all 36 age-sex band rasters",
        f"Mid-year estimate {CFG['census_ref_year']} - update if a better source is available",
        "Positive = WorldPop overestimates",
        "WorldPop typically \u00b15\u201315% vs census for Vietnamese provinces",
    ],
})

add_sheet(wb, "A4_Census_Validation", ext_val_df)

# 2.6. A5. Demographic proportions on the raster grid
if len(age_tifs) > 0 and not np.isnan(raster_total_pop):
    female_tifs = [f for f in age_tifs if f.name.startswith("vnm_f_")]
    male_tifs = [f for f in age_tifs if f.name.startswith("vnm_m_")]

    def stack_sum(files):
        total = None
        for f in files:
            with rasterio.open(f) as src:
                arr = src.read(1).astype("float64")
                nodata = src.nodata
            if nodata is not None:
                arr[arr == nodata] = np.nan
            total = arr if total is None else total + arr
        return total

    r_female = stack_sum(female_tifs)
    r_male = stack_sum(male_tifs)
    r_both = r_female + r_male

    pct_female_raster = (r_female / r_both) * 100
    pf_vals = pct_female_raster[~np.isnan(pct_female_raster)]
    pf_vals = pf_vals[np.isfinite(pf_vals) & (pf_vals > 0)]

    raster_grid_df = pd.DataFrame({
        "Variable": [
            "% Female - N valid pixels", "% Female - Mean", "% Female - SD", "% Female - CV (%)",
            "% Female - Min", "% Female - P25", "% Female - Median", "% Female - P75", "% Female - Max",
            "% Female - IQR",
            "% Female - % pixels with value in [49, 51]",
            "% Female - % pixels identical to national ratio",
        ],
        "Value": [
            fmt_int(len(pf_vals)),
            round(float(np.mean(pf_vals)), 4),
            round(float(np.std(pf_vals, ddof=1)), 4),
            round(float(np.std(pf_vals, ddof=1) / np.mean(pf_vals) * 100), 4),
            round(float(np.min(pf_vals)), 4),
            round(float(np.percentile(pf_vals, 25)), 4),
            round(float(np.median(pf_vals)), 4),
            round(float(np.percentile(pf_vals, 75)), 4),
            round(float(np.max(pf_vals)), 4),
            round(float(np.percentile(pf_vals, 75) - np.percentile(pf_vals, 25)), 4),
            f"{round(float(np.mean((pf_vals >= 49) & (pf_vals <= 51))) * 100, 1)}%",
            f"{round(float(np.mean(np.abs(pf_vals - np.median(pf_vals)) < 0.01)) * 100, 1)}%",
        ],
    })

    add_sheet(wb, "A5_Raster_Proportions", raster_grid_df)
else:
    add_sheet(wb, "A5_Raster_Proportions", pd.DataFrame({"Note": ["Skipped - age-sex rasters not available"]}))

# 3. Part B - H3 resolution-9 assessment ----

# 3.1. Load H3 data
wp_total = gpd.read_file(CFG["worldpop_total"]).drop(columns="geometry")
wp_age = gpd.read_file(CFG["worldpop_age_sex"]).drop(columns="geometry")
h3_sf = gpd.read_file(CFG["h3_geom_gpkg"])
h3_sf["geometry"] = h3_sf.geometry.make_valid()
h3_sf = h3_sf.to_crs(CFG["crs_proj"])
if "h3_id" not in h3_sf.columns:
    id_col = [c for c in h3_sf.columns if c != "geometry"][0]
    h3_sf = h3_sf.rename(columns={id_col: "h3_id"})

# 3.2. B1. Aggregation fidelity
h3_pop_sum = wp_total["total_pop"].sum(skipna=True)
n_zero_hex = (wp_total["total_pop"] == 0).sum()
n_na_hex = wp_total["total_pop"].isna().sum()

agg_df = pd.DataFrame({
    "Metric": [
        "Total H3 hexagons", "H3 hexagons with population > 0", "H3 hexagons with population = 0",
        "H3 hexagons with NA", "Sum of H3 total_pop", "Raster sum (A2)",
        "Aggregation difference", "Aggregation difference (%)",
        "Mean area per hexagon (km\u00b2)", "H3 res-9 nominal area (km\u00b2)",
    ],
    "Value": [
        fmt_int(len(wp_total)),
        fmt_int((wp_total["total_pop"] > 0).sum()),
        fmt_int(n_zero_hex),
        fmt_int(n_na_hex),
        fmt_int(h3_pop_sum),
        fmt_int(raster_total_pop) if not np.isnan(raster_total_pop) else "N/A",
        fmt_int(h3_pop_sum - raster_total_pop) if not np.isnan(raster_total_pop) else "N/A",
        f"{round((h3_pop_sum - raster_total_pop) / raster_total_pop * 100, 3)}%"
        if not np.isnan(raster_total_pop) else "N/A",
        round(h3_sf.geometry.area.mean() / 1e6, 5),
        "~0.1052",
    ],
})

add_sheet(wb, "B1_Aggregation_Fidelity", agg_df, HEADER_STYLE_B)

# 3.3. B2. Variance statistics
age_bands_b = {
    "band": ["0_1", "1_4", "5_9", "10_14", "15_19", "20_24", "25_29", "30_34",
             "35_39", "40_44", "45_49", "50_54", "55_59", "60_64",
             "65_69", "70_74", "75_79", "80_plus"],
    "mid": [0.5, 2.5, 7, 12, 17, 22, 27, 32, 37, 42, 47, 52, 57, 62, 67, 72, 77, 82],
}


def band_sum(row, prefix, bands):
    return sum(row.get(f"{prefix}_{b}", 0) or 0 for b in bands if pd.notna(row.get(f"{prefix}_{b}")))


wp_vars = wp_age.merge(wp_total, on="h3_id", how="left")

wp_vars["total_female"] = wp_vars.apply(lambda r: band_sum(r, "f", age_bands_b["band"]), axis=1)
wp_vars["total_male"] = wp_vars.apply(lambda r: band_sum(r, "m", age_bands_b["band"]), axis=1)
wp_vars["total_both"] = wp_vars["total_female"] + wp_vars["total_male"]
wp_vars["pct_female"] = np.where(
    wp_vars["total_both"] > 0, wp_vars["total_female"] / wp_vars["total_both"] * 100, np.nan
)

demo_band_groups = {
    "pct_child": ["0_1", "1_4", "5_9", "10_14"],
    "pct_youth": ["15_19", "20_24"],
    "pct_millennial": ["25_29", "30_34", "35_39"],
    "pct_genx": ["40_44", "45_49", "50_54"],
    "pct_boomer": ["55_59", "60_64"],
    "pct_elderly": ["65_69", "70_74", "75_79", "80_plus"],
}

for col, bands in demo_band_groups.items():
    grp_sum = wp_vars.apply(lambda r: band_sum(r, "f", bands) + band_sum(r, "m", bands), axis=1)
    wp_vars[col] = np.where(wp_vars["total_both"] > 0, grp_sum / wp_vars["total_both"] * 100, np.nan)


def weighted_median_age(row):
    if row["total_both"] <= 0:
        return np.nan
    bt = np.array([
        (row.get(f"f_{b}", 0) or 0) + (row.get(f"m_{b}", 0) or 0)
        for b in age_bands_b["band"]
    ])
    mids = np.array(age_bands_b["mid"])
    return float(np.sum(mids * bt) / np.sum(bt)) if np.sum(bt) > 0 else np.nan


wp_vars["median_age"] = wp_vars.apply(weighted_median_age, axis=1)
wp_vars = wp_vars[wp_vars["total_pop"].notna() & (wp_vars["total_pop"] > 0)].copy()

vars_list = [
    ("pct_female", "% Female", "%"),
    ("pct_child", "% Children (0\u201314)", "%"),
    ("pct_youth", "% Youth (15\u201324)", "%"),
    ("pct_millennial", "% Millennials (25\u201339)", "%"),
    ("pct_genx", "% Gen X (40\u201354)", "%"),
    ("pct_boomer", "% Baby Boomers (55\u201364)", "%"),
    ("pct_elderly", "% Elderly (65+)", "%"),
    ("median_age", "Median Age", "years"),
    ("total_pop", "Total Population", "persons"),
]

variance_rows = []
for col, label, unit in vars_list:
    x = wp_vars[col].dropna()
    iqr = float(np.percentile(x, 75) - np.percentile(x, 25))
    variance_rows.append({
        "Variable": label, "Unit": unit, "N_hexagons": len(x),
        "Mean": round(float(x.mean()), 4), "SD": round(float(x.std(ddof=1)), 4),
        "CV_pct": round(float(x.std(ddof=1) / x.mean() * 100), 3),
        "Min": round(float(x.min()), 4),
        "P25": round(float(np.percentile(x, 25)), 4),
        "Median": round(float(x.median()), 4),
        "P75": round(float(np.percentile(x, 75)), 4),
        "Max": round(float(x.max()), 4),
        "IQR": round(iqr, 5),
        "IQR_zero": "YES \u2014 uniform" if iqr == 0 else "No",
        "Range_over_mean_pct": round(float((x.max() - x.min()) / x.mean() * 100), 2),
    })
variance_df = pd.DataFrame(variance_rows)

add_sheet(wb, "B2_Variance_Statistics_H3", variance_df, HEADER_STYLE_B)

# 3.4. B3. Moran's I (H3)
demo_cols = ["pct_female", "pct_child", "pct_youth", "pct_millennial",
             "pct_genx", "pct_boomer", "pct_elderly", "median_age", "total_pop"]
demo_labels = ["% Female", "% Children (0\u201314)", "% Youth (15\u201324)",
               "% Millennials (25\u201339)", "% Gen X (40\u201354)",
               "% Baby Boomers (55\u201364)", "% Elderly (65+)",
               "Median Age", "Total Population"]

wp_sf = h3_sf.merge(wp_vars[["h3_id"] + demo_cols], on="h3_id", how="inner")
wp_sf = wp_sf[wp_sf["pct_female"].notna()].reset_index(drop=True)

coords = np.column_stack([wp_sf.geometry.centroid.x, wp_sf.geometry.centroid.y])
knn_w = KNN.from_array(coords, k=CFG["knn_k"])
knn_w.transform = "r"

moran_rows = []
for col, label in zip(demo_cols, demo_labels):
    mt = Moran(wp_sf[col].to_numpy(), knn_w, two_tailed=True)
    z = (mt.I - mt.EI) / np.sqrt(mt.VI_norm)
    sig = "***" if mt.p_sim < 0.001 else "**" if mt.p_sim < 0.01 else "*" if mt.p_sim < 0.05 else "n.s."
    moran_rows.append({
        "Variable": label,
        "Moran_I": round(mt.I, 4),
        "Z_score": round(float(z), 2),
        "p_value": mt.p_sim,
        "Sig": sig,
        "Interpretation": (
            "High Moran's I + high CV \u2192 genuine spatial clustering" if col == "total_pop"
            else "High Moran's I + near-zero CV \u2192 uniform disaggregation artifact"
        ),
    })
moran_df = pd.DataFrame(moran_rows)

add_sheet(wb, "B3_Morans_I_H3", moran_df, HEADER_STYLE_B)

# 3.5. B4. Correlation with accessibility
access_wide = pd.read_csv(CFG["access_wide_csv"])
id_col_candidates = [c for c in ["h3_id", "origin_id", "id"] if c in access_wide.columns]
if id_col_candidates and id_col_candidates[0] != "h3_id":
    access_wide = access_wide.rename(columns={id_col_candidates[0]: "h3_id"})

acc_col_map = {
    "retail_grocery": "retail_grocery_cum30m",
    "education": "education_cum30m",
    "healthcare": "healthcare_cum30m",
    "food_beverage": "food_beverage_cum30m",
    "parks": "parks_public_open_space_cum30m",
}
acc_col_map = {k: v for k, v in acc_col_map.items() if v in access_wide.columns}

corr_base = wp_vars.merge(
    access_wide[["h3_id"] + list(acc_col_map.values())], on="h3_id", how="inner"
)
corr_base = corr_base[corr_base["pct_female"].notna()]

corr_vars = ["pct_female", "pct_child", "pct_youth", "pct_millennial",
             "pct_genx", "pct_boomer", "pct_elderly", "median_age"]
corr_labels = ["% Female", "% Children (0\u201314)", "% Youth (15\u201324)",
               "% Millennials (25\u201339)", "% Gen X (40\u201354)",
               "% Baby Boomers (55\u201364)", "% Elderly (65+)", "Median Age"]

corr_rows = []
for var, label in zip(corr_vars, corr_labels):
    row = {"Variable": label}
    for nm, col_name in acc_col_map.items():
        pair = corr_base[[var, col_name]].dropna()
        row[f"r_{nm}"] = round(float(pair[var].corr(pair[col_name])), 3) if len(pair) > 2 else np.nan
    row["Max_abs_r"] = round(max(abs(v) for v in row.values() if isinstance(v, (int, float)) and not np.isnan(v)), 3)
    corr_rows.append(row)
corr_df = pd.DataFrame(corr_rows)

add_sheet(wb, "B4_Correlation_Accessibility", corr_df, HEADER_STYLE_B)

# 3.6. B5. Key diagnostic numbers for paper
key_df = pd.DataFrame({
    "Level": ["Raster Grid"] * 5 + ["H3 Resolution-9"] * 12,
    "Metric": [
        "Pixel resolution (approx.)", "N valid pixels", "WorldPop total population",
        "Census reference population (2018)", "WorldPop vs census difference (%)",
        "N hexagons total", "N hexagons with pop > 0", "Mean hex area (km\u00b2)",
        "% Female - Mean (%)", "% Female - SD", "% Female - CV (%)", "% Female - IQR",
        "Median Age - Mean (years)", "Median Age - SD", "Median Age - CV (%)",
        "% Female - Moran's I", "Median Age - Moran's I",
    ],
    "Value": [
        f"~{pixel_res_m if CFG['total_pop_tif'].exists() else 'N/A'} m",
        fmt_int(n_valid_pixels) if CFG["total_pop_tif"].exists() else "N/A",
        fmt_int(raster_total_pop) if not np.isnan(raster_total_pop) else "N/A",
        fmt_int(CFG["census_ref_pop"]),
        f"{round((raster_total_pop - CFG['census_ref_pop']) / CFG['census_ref_pop'] * 100, 2)}%"
        if not np.isnan(raster_total_pop) else "N/A",
        fmt_int(len(wp_total)),
        fmt_int(len(wp_vars)),
        round(h3_sf.geometry.area.mean() / 1e6, 5),
        round(float(wp_vars["pct_female"].mean(skipna=True)), 3),
        round(float(wp_vars["pct_female"].std(ddof=1)), 4),
        round(float(wp_vars["pct_female"].std(ddof=1) / wp_vars["pct_female"].mean() * 100), 3),
        round(float(np.percentile(wp_vars["pct_female"].dropna(), 75) - np.percentile(wp_vars["pct_female"].dropna(), 25)), 5),
        round(float(wp_vars["median_age"].mean(skipna=True)), 2),
        round(float(wp_vars["median_age"].std(ddof=1)), 3),
        round(float(wp_vars["median_age"].std(ddof=1) / wp_vars["median_age"].mean() * 100), 3),
        round(float(moran_df.loc[moran_df["Variable"] == "% Female", "Moran_I"].iloc[0]), 4),
        round(float(moran_df.loc[moran_df["Variable"] == "Median Age", "Moran_I"].iloc[0]), 4),
    ],
})

add_sheet(wb, "B5_Key_Numbers_Paper", key_df, HEADER_STYLE_B)

# 4. Save ----
wb.save(CFG["out_excel"])
